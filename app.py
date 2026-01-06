from flask import Flask, request, render_template, redirect, url_for, flash, send_from_directory
from werkzeug.utils import secure_filename
from io import StringIO
import os
import pandas as pd
from datetime import datetime
from dateutil import parser as date_parser
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv'}
LOG_FOLDER = 'logs'

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = os.environ.get('FLASK_SECRET', 'dev_secret_key')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LOG_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def find_columns(df):
    # Try to detect likely columns: date, time, quantity, price, product
    columns = { 'date': None, 'time': None, 'quantity': None, 'price': None, 'product': None }
    lower_cols = {c.lower(): c for c in df.columns}

    # Mapping rules
    date_candidates = ['date', 'transaction date', 'trade date', 'trade_date', 'timestamp', 'datetime']
    time_candidates = ['time', 'trade time', 'timestamp', 'datetime', 'trade_time']
    qty_candidates = ['qty', 'quantity', 'amount', 'shares', 'share']
    # Prefer the 'total' column if present, otherwise fall back to unit price
    total_candidates = ['total', 'total value', 'total_amount', 'total_amount', 'value', 'proceeds_total']
    price_candidates = ['price', 'unit price', 'unit_price', 'proceeds', 'amount per share']
    product_candidates = ['product', 'symbol', 'ticker', 'asset', 'security', 'stock']

    for cand in date_candidates:
        for k, original in lower_cols.items():
            if cand in k or cand.replace(' ', '_') in k:
                columns['date'] = original
                break
        if columns['date']:
            break

    for cand in time_candidates:
        for k, original in lower_cols.items():
            if cand in k or cand.replace(' ', '_') in k:
                columns['time'] = original
                break
        if columns['time']:
            break

    for cand in qty_candidates:
        for k, original in lower_cols.items():
            if cand in k or cand.replace(' ', '_') in k:
                columns['quantity'] = original
                break
        if columns['quantity']:
            break

    # First look for a 'total' style column and prefer it as the price candidate
    for cand in total_candidates:
        for k, original in lower_cols.items():
            if cand in k or cand.replace(' ', '_') in k:
                columns['price'] = original
                break
        if columns['price']:
            break

    # If we didn't find a total column, look for unit price
    if not columns['price']:
        for cand in price_candidates:
            for k, original in lower_cols.items():
                if cand in k or cand.replace(' ', '_') in k:
                    columns['price'] = original
                    break
            if columns['price']:
                break

    for cand in product_candidates:
        for k, original in lower_cols.items():
            if cand in k or cand.replace(' ', '_') in k:
                columns['product'] = original
                break
        if columns['product']:
            break

    return columns


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/download/<path:filename>')
def download(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(url_for('index'))
    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(url_for('index'))
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        # We'll read directly into pandas without saving to disk
        content = file.read().decode('utf-8')
        try:
            df = pd.read_csv(StringIO(content))
        except Exception as e:
            flash('Failed to read CSV file: ' + str(e))
            return redirect(url_for('index'))

        # Detect columns; pass to mapping page
        suggestions = find_columns(df)
        request_id = datetime.now().strftime('%Y%m%d%H%M%S%f')
        csv_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{request_id}.csv")
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write(content)

        return render_template('map_columns.html', columns=list(df.columns), suggestions=suggestions, csv_path=csv_path)
    else:
        flash('Invalid file type. Only CSV allowed')
        return redirect(url_for('index'))


def parse_date(value):
    if pd.isna(value):
        return None
    try:
        # First try exact dd/mm/YYYY
        v = str(value).strip()
        try:
            dt = datetime.strptime(v, "%d/%m/%Y")
            return dt.date()
        except Exception:
            pass
        # Fall back to dateutil with dayfirst=True to prefer dd/mm parsing
        dt = date_parser.parse(v, dayfirst=True)
        # convert to date only (no time component)
        return dt.date()
    except Exception:
        return None


def parse_datetime(date_value, time_value=None):
    # Combine date and optional time values into a datetime object (or return None)
    if pd.isna(date_value):
        return None
    try:
        dstr = str(date_value).strip()
        if time_value is not None and not pd.isna(time_value) and str(time_value).strip() != '':
            tstr = str(time_value).strip()
            combined = f"{dstr} {tstr}"
            dt = date_parser.parse(combined, dayfirst=True)
            return dt
        # Parse date only
        try:
            dd = datetime.strptime(dstr, "%d/%m/%Y")
            return datetime(dd.year, dd.month, dd.day)
        except Exception:
            return date_parser.parse(dstr, dayfirst=True)
    except Exception:
        return None


def format_date_for_display(value):
    if value is None:
        return ''
    try:
        if isinstance(value, datetime):
            # if has time component, show YYYY-MM-DD HH:MM, else just date
            if value.time() and (value.hour != 0 or value.minute != 0 or value.second != 0):
                return value.strftime('%Y-%m-%d %H:%M')
            return value.strftime('%Y-%m-%d')
        # If it's a date (not datetime)
        try:
            return value.strftime('%Y-%m-%d')
        except Exception:
            # Fallback: try to parse string and format
            dt = date_parser.parse(str(value), dayfirst=True)
            if hasattr(dt, 'time') and (dt.hour != 0 or dt.minute != 0):
                return dt.strftime('%Y-%m-%d %H:%M')
            return dt.strftime('%Y-%m-%d')
    except Exception:
        return str(value)


@app.route('/calculate', methods=['POST'])
def calculate():
    csv_path = request.form.get('csv_path')
    date_col = request.form.get('date_col')
    qty_col = request.form.get('qty_col')
    price_col = request.form.get('price_col')
    product_col = request.form.get('product_col')
    time_col = request.form.get('time_col')

    if not csv_path or not os.path.exists(csv_path):
        flash('CSV file not found. Re-upload file.')
        return redirect(url_for('index'))

    df = pd.read_csv(csv_path)

    # Validate columns
    for c in [date_col, qty_col, price_col, product_col]:
        if not c or c not in df.columns:
            flash(f'Missing column: {c}')
            return redirect(url_for('index'))

    # Normalize data - include optional time column if provided
    cols_to_extract = [date_col, qty_col, price_col, product_col]
    if time_col and time_col in df.columns:
        cols_to_extract.append(time_col)
    df = df[cols_to_extract]
    if time_col and time_col in df.columns:
        df.columns = ['date', 'quantity', 'price', 'product', 'time']
    else:
        df.columns = ['date', 'quantity', 'price', 'product']

    # Parse dates and combine with optional time to produce datetimes
    if 'time' in df.columns:
        df['date'] = df.apply(lambda r: parse_datetime(r['date'], r['time']), axis=1)
    else:
        df['date'] = df['date'].apply(parse_date)
    # Preserve original CSV order to resolve same-day ties and ensure deterministic FIFO
    df['_row_id'] = df.index.astype(int)
    df = df.sort_values(['date', '_row_id'])
    # Strip commas and spaces from quantities (e.g. "2,000" -> 2000) before numeric conversion
    df['quantity'] = df['quantity'].astype(str).str.replace(',', '', regex=False).str.replace(' ', '', regex=False)
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce')
    # Strip commas and spaces from prices as well (e.g. "1,234.56" -> 1234.56)
    df['price'] = df['price'].astype(str).str.replace(',', '', regex=False).str.replace(' ', '', regex=False)
    df['price'] = pd.to_numeric(df['price'], errors='coerce')
    df['price'] = pd.to_numeric(df['price'], errors='coerce')

    # Determine buy vs sell
    # Only use the sign of the quantity to determine buy (positive) or sell (negative)
    df['sign'] = df['quantity'].apply(lambda q: 1 if q >= 0 else -1)

    # Normalize positive quantities for both sides, use sign separately
    df['abs_qty'] = df['quantity'].abs()

    # Determine if the selected price column represents a transaction total
    price_col_lower = price_col.lower() if price_col else ''
    price_is_total = 'total' in price_col_lower

    # Compute unit_price: if price_is_total, divide total by absolute qty, else use price directly
    if price_is_total:
        # Avoid division by zero by using NaN for rows with zero abs_qty
        df['unit_price'] = df.apply(lambda r: (abs(r['price']) / r['abs_qty']) if pd.notna(r['abs_qty']) and r['abs_qty'] != 0 else pd.NA, axis=1)
        # Exclude transactions where the selected price column is a transaction total and the total value is zero
        zero_total_mask = df['price'].fillna(0).abs() == 0
        if zero_total_mask.any():
            # Drop zero-total transactions entirely (they should not be used for FIFO matching)
            df = df.loc[~zero_total_mask].copy()
    else:
        df['unit_price'] = df['price']

    # Normalize product values and process per-product FIFO matching and build matrices
    df['product'] = df['product'].astype(str).str.strip().str.upper()
    product_reports = []
    grand_total_gain = 0.0
    all_years = set()

    for product, group in df.groupby('product'):
        # group is already sorted by date,_row_id, but sort again for safety
        g = group.sort_values(['date', '_row_id']).reset_index(drop=True)
        buys = []  # each: dict{id, date, qty, price, remaining}
        sells = []  # each: dict{id, date, qty, price}
        matches = {}  # (buy_id, sell_id) -> qty matched
        sale_profits = []
        total_gain = 0.0
        # open a per-product section in the log
        log_lines = []
        log_lines.append(f"Product: {product}")
        log_lines.append("--- Transactions ---")
        for _, row in g.iterrows():
            row_date = row['date']
            qty = float(row['abs_qty'])
            price = float(row['unit_price']) if pd.notna(row['unit_price']) else 0.0
            sign = int(row['sign'])

            if sign > 0:
                buy_id = len(buys)
                # convert date to ISO string for display and sorting
                buy_entry = {'id': buy_id, 'date': format_date_for_display(row_date), 'date_raw': row_date, 'qty': qty, 'price': price, 'remaining': qty}
                buys.append(buy_entry)
                log_lines.append(f"BUY id={buy_id} date={buy_entry['date']} qty={qty} price={price}")
            else:
                sell_id = len(sells)
                sell_entry = {'id': sell_id, 'date': format_date_for_display(row_date), 'date_raw': row_date, 'qty': qty, 'price': price}
                sells.append(sell_entry)
                log_lines.append(f"SELL id={sell_id} date={sell_entry['date']} qty={qty} price={price}")
                to_sell = qty
                realized_for_sale = 0.0
                # match FIFO to buys
                b_index = 0
                while to_sell > 1e-8 and b_index < len(buys):
                    lot = buys[b_index]
                    if lot['remaining'] <= 1e-8:
                        b_index += 1
                        continue
                    take = min(lot['remaining'], to_sell)
                    matches[(lot['id'], sell_id)] = matches.get((lot['id'], sell_id), 0.0) + take
                    realized = (price - lot['price']) * take
                    realized_for_sale += realized
                    total_gain += realized
                    lot['remaining'] -= take
                    to_sell -= take
                    log_lines.append(f"  MATCH sell_id={sell_id} buy_id={lot['id']} take={take} buy_price={lot['price']} sell_price={price} realized={realized}")
                    if lot['remaining'] <= 1e-8:
                        b_index += 1

                if to_sell > 1e-8:
                    # Unmatched sell quantity - record as unmatched with None profit contribution
                    # We simply record the unmatched amount as a match from a virtual None buy (not represented in matrix)
                    matches[(None, sell_id)] = matches.get((None, sell_id), 0.0) + to_sell
                    log_lines.append(f"  UNMATCHED sell_id={sell_id} qty={to_sell}")
                    # realized_for_sale for unmatched remains unchanged (can't compute cost)

                sale_profits.append({'sell_id': sell_id, 'profit': realized_for_sale, 'date': format_date_for_display(row_date), 'price': price})

        # Build matrix rows x cols containing matched quantities
        rows = len(buys)
        cols = len(sells)
        matrix = [[0.0 for _ in range(cols)] for _ in range(rows)]
        for (b, s), q in matches.items():
            if b is None:
                continue
            matrix[b][s] = q

        # Add matrix and summary to log lines
        log_lines.append("--- Matrix (rows=buys, cols=sells) ---")
        for b_idx, b in enumerate(buys):
            rowvals = [str(int(matrix[b_idx][s_idx])) if b_idx < len(matrix) and s_idx < len(matrix[b_idx]) else '0' for s_idx in range(cols)]
            log_lines.append(f"BUY {b_idx} -> " + ",".join(rowvals))
        log_lines.append("--- Sale profits ---")
        for sp in sale_profits:
            log_lines.append(f"SALE id={sp['sell_id']} date={sp['date']} profit={sp['profit']}")

        # Write the log section for this product to a per-run log file (append)
        base = os.path.splitext(os.path.basename(csv_path))[0]
        log_path = os.path.join(LOG_FOLDER, f"{base}.log")
        with open(log_path, 'a', encoding='utf-8') as lf:
            lf.write("\n".join(log_lines) + "\n\n")

        # Compute summary info for this product
        buys_count = len(buys)
        sells_count = len(sells)
        # first and last transaction dates (skip None)
        dates = [d for d in g['date'].tolist() if d is not None]
        first_date = min(dates) if dates else None
        last_date = max(dates) if dates else None
        remaining_qty = sum([b.get('remaining', 0.0) for b in buys])
        # Mark as irregular if any unit_price for this product is exactly zero
        # Use 'unit_price' if present in the group (computed earlier when total column chosen), otherwise use 'price'
        unit_price_col = 'unit_price' if 'unit_price' in g.columns else 'price'
        irregular = False
        try:
            irregular = ((g[unit_price_col].fillna(0) == 0) & (g['abs_qty'] > 0)).any()
        except Exception:
            irregular = False
        # Profit by year
        profit_by_year = {}
        for sp in sale_profits:
            dstr = sp.get('date')
            if not dstr:
                continue
            try:
                dt = date_parser.parse(dstr, dayfirst=True)
                y = dt.year
            except Exception:
                continue
            profit_by_year[y] = profit_by_year.get(y, 0.0) + sp.get('profit', 0.0)
            all_years.add(y)

        product_reports.append({'product': product, 'buys': buys, 'sells': sells, 'matrix': matrix, 'sale_profits': sale_profits, 'total_gain': total_gain, 'summary': {'buys_count': buys_count, 'sells_count': sells_count, 'first_date': first_date, 'last_date': last_date, 'first_date_str': format_date_for_display(first_date) if first_date is not None else '', 'last_date_str': format_date_for_display(last_date) if last_date is not None else '', 'remaining_qty': remaining_qty, 'profit_by_year': profit_by_year, 'irregular': irregular}})
        grand_total_gain += total_gain

    years = sorted(list(all_years))
    # Attempt to create an Excel workbook with one sheet per product and a summary sheet
    base = os.path.splitext(os.path.basename(csv_path))[0]
    xlsx_name = f"{base}.xlsx"
    xlsx_path = os.path.join(app.config['UPLOAD_FOLDER'], xlsx_name)
    try:
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_rows = []
            for pr in product_reports:
                row = {'Product': pr['product']}
                for y in years:
                    row[str(y)] = pr['summary']['profit_by_year'].get(y, 0.0)
                row['Total Buys'] = pr['summary']['buys_count']
                row['Total Sells'] = pr['summary']['sells_count']
                row['First Transaction'] = pr['summary'].get('first_date_str', '')
                row['Last Transaction'] = pr['summary'].get('last_date_str', '')
                row['Remaining Qty'] = pr['summary']['remaining_qty']
                summary_rows.append(row)
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Per-product sheets: single FIFO matrix with buy rows and sell columns
            for pr in product_reports:
                sheet = pr['product'][:31] if pr['product'] else 'PRODUCT'
                rows = len(pr['buys'])
                cols = len(pr['sells'])
                # Create DataFrame with 3 header rows and 3 header columns
                # If sells present, reserve two extra rows: one blank row and one for per-column profit totals
                nrows = 3 + rows + (2 if cols > 0 else 0)
                ncols = 3 + cols
                table = [["" for _ in range(ncols)] for _ in range(nrows)]

                # Fill sell headers across top 3 rows
                for s_idx, s in enumerate(pr['sells']):
                    c = 3 + s_idx
                    table[0][c] = s['date']
                    table[1][c] = int(s['qty'])
                    table[2][c] = "{:.2f}".format(s['price'])

                # Fill buy headers down first 3 columns
                for b_idx, b in enumerate(pr['buys']):
                    r = 3 + b_idx
                    table[r][0] = b['date']
                    table[r][1] = int(b['qty'])
                    table[r][2] = "{:.2f}".format(b['price'])

                # Fill matrix body
                for b_idx in range(rows):
                    for s_idx in range(cols):
                        v = pr['matrix'][b_idx][s_idx]
                        table[3 + b_idx][3 + s_idx] = int(v) if v and v != 0 else ''

                # Compute total profit per sell column and write beneath the matrix
                # Reserve an extra blank row above the profit row for spacing
                # Profit for a column = sum_over_buys( matched_qty * (sell_price - buy_price) )
                if cols > 0:
                    profit_row_idx = 3 + rows + 1  # 0-based index into table for the profit row (one blank row first)
                    # Label for the profit row in first column
                    table[profit_row_idx][0] = 'Column Profit'
                    col_profits = [0.0 for _ in range(cols)]
                    for s_idx in range(cols):
                        s_price = pr['sells'][s_idx]['price']
                        total = 0.0
                        for b_idx in range(rows):
                            qty = pr['matrix'][b_idx][s_idx]
                            if qty and qty != 0:
                                b_price = pr['buys'][b_idx]['price']
                                total += qty * (s_price - b_price)
                        col_profits[s_idx] = total
                        # place numeric profit value into table (Excel will store as number)
                        table[profit_row_idx][3 + s_idx] = round(total, 2)

                matrix_df = pd.DataFrame(table)
                matrix_df.to_excel(writer, sheet_name=sheet, index=False, header=False)
                # Post-process sheet with openpyxl to apply UX formatting
                ws = writer.sheets[sheet]
                bold = Font(bold=True)
                center = Alignment(horizontal='center', vertical='center')
                header_fill = PatternFill(start_color='FFEEEEEE', end_color='FFEEEEEE', fill_type='solid')

                ws['A3'] = 'date (purchases)'
                ws['B3'] = 'quantity (purchases)'
                ws['C1'] = 'date (sales)'
                ws['C2'] = 'quantity (sales)'
                ws['C3'] = 'price (sales)'
                
                # Apply black background and white bold font to label cells
                black_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                white_bold = Font(bold=True, color='FFFFFFFF')
                for coord in [('A', 3), ('B', 3), ('C', 1), ('C', 2), ('C', 3)]:
                    cell = ws[f"{coord[0]}{coord[1]}"]
                    cell.font = white_bold
                    cell.fill = black_fill
                    cell.alignment = center

                # Add Purchases and Sales labels with arrows
                ws['B1'] = 'Purchases →'
                ws['A2'] = '\u2190 Sales'

                rows = len(pr['buys'])
                cols = len(pr['sells'])

                # Style the Purchases and Sales label cells (B1 and A2)
                for coord in [('B', 1), ('A', 2)]:
                    cell = ws[f"{coord[0]}{coord[1]}"]
                    cell.font = bold
                    cell.alignment = center

                # Apply bold + center to first 3 columns for buy headers (rows 4..)
                for r in range(1, 4):
                    for c in range(4, 4 + cols):
                        cell = ws.cell(row=r, column=c)
                        cell.font = bold
                        cell.alignment = center
                        cell.fill = header_fill
                        # qty column (col 2) integer, price column (col 3) currency
                        if c == 2:
                            cell.number_format = '0'
                        if c == 3:
                            cell.number_format = '#,##0.00'

                # Apply bold + center to first 3 columns for buy headers (rows 4..)
                for c in range(1, 4):
                    for r in range(4, 4 + rows):
                        cell = ws.cell(row=r, column=c)
                        cell.font = bold
                        cell.alignment = center
                        cell.fill = header_fill
                        # qty column (col 2) integer, price column (col 3) currency
                        if c == 2:
                            cell.number_format = '0'
                        if c == 3:
                            cell.number_format = '#,##0.00'

                # Format matrix body as integers
                for r in range(4, 4 + rows):
                    for c in range(4, 4 + cols):
                        cell = ws.cell(row=r, column=c)
                        cell.number_format = '0'

                # Format and label the profit row beneath the matrix if present
                if cols > 0:
                    # profit_row now after an extra blank row: 1-based index
                    profit_row = 5 + rows
                    # Label in column A
                    cell = ws.cell(row=profit_row, column=1)
                    cell.value = 'Column Profit'
                    cell.font = bold
                    cell.alignment = center
                    # Apply number format to profit cells under each sell column
                    for c in range(4, 4 + cols):
                        pcell = ws.cell(row=profit_row, column=c)
                        pcell.number_format = '#,##0.00'
                        pcell.font = bold
                        pcell.alignment = center

                # No freeze panes
                # Adjust column widths
                for col_idx in range(1, 4 + cols):
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    if col_idx <= 3:
                        ws.column_dimensions[col_letter].width = 18
                    else:
                        ws.column_dimensions[col_letter].width = 12
    except Exception:
        # If Excel generation fails, continue without crashing; the link will still be shown
        pass

    return render_template('results.html', product_reports=product_reports, total_gain=grand_total_gain, years=years, xlsx_name=xlsx_name)


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)

